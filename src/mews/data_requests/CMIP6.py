"""
CMIP6 data collection and calculations. Used when accounting for global warming
in MEWS.

Calculates temperature change values from a baseline year to a future year
for a specific latitude-longitude input. Compiles information from historical
global climate models and 5 SSP scenario models.

Created on Wed Jun 20 13:00:31 2022

@author: tschoste with updates after summer 2022 by dlvilla

Copyright Notice
=================

Copyright 2023 National Technology and Engineering Solutions of Sandia, LLC. 
Under the terms of Contract DE-NA0003525, there is a non-exclusive license 
for use of this work by or on behalf of the U.S. Government. 
Export of this program may require a license from the 
United States Government.

Please refer to the LICENSE.md file for a full description of the license
terms for MEWS. 

The license for MEWS is the Modified BSD License and copyright information
must be replicated in any derivative works that use the source code.

----

"""

import openpyxl
import os
import urllib.request
import warnings
warnings.filterwarnings(action='ignore',category=UserWarning)
# Ignore this warning. The code works anyway and cfgrib is not needed for the cases
# use by MEWS
#/home/dlvilla/python/miniconda38_10/lib/python3.8/site-packages/xarray/backends/cfgrib_.py:29: 
# UserWarning: Failed to load cfgrib - most likely there is a problem accessing the ecCodes library. 
# Try `import cfgrib` to get the full error message
#  warnings.warn(

import xarray as xr
warnings.filterwarnings(action='default',category=UserWarning)
import numpy as np
import math
import matplotlib.pyplot as plt
import scipy.stats as st
from scipy.optimize import curve_fit
import multiprocessing as mp
import threading as thr
import logging
import pandas as pd

from mews.utilities.utilities import filter_cpu_count

########## Initial Startup ##############
warnings.simplefilter(action='ignore') #Removes depreciation warnings


class CMIP_Data(object):
    """
    
    >>> obj = CMIP_Data(lat_desired,
                        lon_desired,
                        baseline_year,
                        end_year,
                        model_guide,
                        file_path,
                        data_folder,
                        )
    
    
    Results can be found in obj.delT which returns a dictionary of the 
    temperature change values for each scenario based on end_year.
    
    More complete results can be found in obj.total_model_data. This is a 
    dictionary seperated based on scenario, that contains the following
    information based on the latitude-longitude input.
    
    .. list-table::
       :widths: 25 75
       :header-rows: 1

       * - Attribute
         - Description
       * - .start_year
         - The first year of data available for the scenario
       * - .end_year
         - The last year of data available for the scenario
       * - .year_temp_dict
         - A dictionary of lists containing the temperature values for each model
       * - .scenario
         - The selected scenario
       * - .avg_error_list
         - A list of the normalized interpolation error values for each model
       * - .delT_list
         - A list of all averaged temperature change values with the first entry 
           corresponding to the start_year and the last entry corresponding to the 
           end_year
       * - .CI_list
         - A list of the 95% confidence interval bounds for delT_list
       * - .delT_list_reg
         - A list of all regressed temperature change values with the first entry
           corresponding to the start_year and the last entry corresponding to the
           end_year
       * - .avg_error
         - The averaged error value for every model used in the scenario
         
    Example
    -------
    
    >>> obj.total_model_data["SSP119"].start_year
    2015
    
    Plots can be generated through obj.results1 and obj.results2 (see below).
    
    
    Parameters
    ----------
    lat_desired : float
        Latitude value used for temperature change calculations. Must be
        between -90 and 90 [degrees].
    
    lon_desired : float
        Longitude value used for temperature change calculations. Must be 
        between -360 (when using degrees W) and 360 (when using degrees E).
        
    baseline_year : int
        Year input for the baseline temperature change. All delta_T's will 
        be calculated from this year. 1850 <= year <= 2014.
        
    end_year : int
        Future year input for which all temperature change values will be 
        calculated to. 2015 <= year <= 2100.
        
    
    model_guide : str : optional : Default = None
        Filename and path for excel spreadsheet which contains the names and links 
        to all models used in the calculations. If None, then the default
        model guide created for MEWS is used.
    
    data_folder : str - RELATIVE FOLDER PATH : optional : Default = None 
        The name of the folder where all of the CMIP data files will be 
        downloaded to a directory "CMIP6_Data_Files" at the same level as
        the MEWS repository so that files are not included in the MEWS.git folder
        structure.
        
    world_map : bool : optional : Default = True
        Displays the latitude-longitude location desired on a world map.
        Can be useful in double checking the location is selected correctly.
        
    calculate_error : bool : optional : Default = True
        If true, will calculate normalization and distance error for all
        models. Sighlty faster to run if set to False.
        
    display_logging : bool : optional : Default = False
        If set to True, the logging statements will be printed to the console
        as well as saved to a file in the current directory.
        
    display_plots : bool : optional : Default = True
        If set to False, the computation for the plotting will be done, but
        no plots will be shown. Used in unittesting.
        
    run_parallel : bool : optional : Default = True
        If True, asyncronous paralellization will be used to speed up calculations
        accross the various scenarios.
    
    proxy : str : optional : Default = None
        Must indicate the proxy being used for downloads if a proxy server is
        being used.
        example: https://proxy.institution.com:8080
        
    gcm_to_skip : list of str : optional : Default = []
        allows entry of GCM names that will be skipped. If a GCM server will not
        download, then this can be used to skip inclusion of that GCM.
        
    polynomial_fit_order : dict, optional : Default = _default_poly_order_ (string below)
        controls what order polynomial is fit to each scenario.
        
    align_gcm_to_historical : bool, optional : Default = False
        False : allows there to be a discontinuity between historic GCM ensemble
                and the future GCM projections. There are significant differences
                in the scenario start points though.
        True : shifts all GCM data for each scenario to start at the historic
               average for the baseline year. GCM data is shifted up or down
               to accomplish this.
    
    num_cpu : int or None : optional : Default = None
        None - use the maximum number of cpu's available minus one
        int - use the requested number of cpu's as long as it is less than
              the number of cpu's available minus one.
              Only applies when run_parallel = True
    
    Returns
    -------
    None 
    
    """
    # constants 
    _valid_scenarios = ["historical","SSP119","SSP126","SSP245","SSP370","SSP585"]
    _cmip6_start_year = 1850
    _cmip6_baseline_end_year = 1900
    _historical_scen_str = _valid_scenarios[0]
    _default_poly_order = {_valid_scenarios[0]:6,
                           _valid_scenarios[1]:3,
                           _valid_scenarios[2]:2,
                           _valid_scenarios[3]:2,
                           _valid_scenarios[4]:2,
                           _valid_scenarios[5]:2,}
    
    
    def __init__(self, lat_desired, 
                 lon_desired,
                 baseline_year,
                 end_year,
                 model_guide,
                 file_path,
                 data_folder=None,
                 scenario_list=_valid_scenarios,
                 world_map=False,
                 calculate_error=True,
                 display_logging=False,
                 display_plots=True,
                 run_parallel=True,
                 proxy=None,
                 gcm_to_skip=[],
                 polynomial_fit_order=_default_poly_order,
                 align_gcm_to_historical=False,
                 num_cpu=None):
        
        self._num_cpu = filter_cpu_count(num_cpu)
        
        if not proxy is None:
            os.environ['http_proxy'] = proxy 
            os.environ['HTTP_PROXY'] = proxy
            os.environ['https_proxy'] = proxy
            os.environ['HTTPS_PROXY'] = proxy
            
        self.world_map = world_map
        self.lat_desired = lat_desired
        self.lon_desired = lon_desired
        self.baseline_year = baseline_year
        self.end_year = end_year
        self.calculate_error = calculate_error
        self.model_guide = model_guide
        self.data_folder = data_folder
        if not self._historical_scen_str in scenario_list:
            scenario_list.insert(0,"historical")
        self.scenario_list = scenario_list
        self.display_plots = display_plots
        self.data_folder = data_folder
        self.gcm_to_skip = gcm_to_skip
        self._polynomial_fit_order = polynomial_fit_order
        self._Kelvin_to_Celcius_offset = 273.15
        self._align_gcm_to_historical = align_gcm_to_historical
        
        self._check_lat_lon(lat_desired, lon_desired)
        
        #Initializing logging
        global logger
        logger = logging.getLogger(__name__)
        logger.setLevel(logging.INFO)
        formatter = logging.Formatter("%(asctime)s | %(levelname)s | %(message)s")
        logger.handlers = []

        file_handler = logging.FileHandler("CMIP6_Data_Log.log")
        file_handler.setFormatter(formatter)
        logger.addHandler(file_handler)
        logger.propagate = False
        if display_logging:
            stream_handler = logging.StreamHandler()
            stream_handler.setFormatter(formatter)
            logger.addHandler(stream_handler)
            logger.propagate = False
        logger.info("Script Initiation")
        
        #Testing Cartopy module import
        try:
            global ccrs
            global cf
            import cartopy.crs as ccrs
            import cartopy.feature as cf
            self.use_cartopy = True
            #urllib.request.urlretrieve("https://testingcaseinwhichthisfails")
        except:
            logger.warning("Unable to import cartopy")
            self.use_cartopy = False
        

        
        self._CMIP_Data_Collection(run_parallel)
        self.run_parallel = run_parallel
    
    def _check_lat_lon(self,lat_desired,lon_desired):
        #Checking latitude and longitude inputs
        if self.lon_desired < 0: #If negative lon is inputted
            self.lon_desired += 360 
        if self.lon_desired > 360 or self.lon_desired < 0:
            raise ValueError("Longitude input is out of range")
        if self.lat_desired > 90 or self.lat_desired < -90:
            raise ValueError("Latitude input is out of range")
    
    def new_lat_lon(self,lat_desired,lon_desired):
        """
        TODO - NOT TESTED verify that this gets the same answer if you run an analysis
        independently twice.
        
        Performs the exact same analysis as on initialization but 
        for a new lat/lon 

        Parameters
        ----------
        lat_desired : float
            latitude for new analysis
        lon_desired : TYPE
            longitude for new analysis

        Returns
        -------
        None.

        """
        self._check_lat_lon(lat_desired,lon_desired)
        
        self.lat_desired = lat_desired
        self.lon_desired = lon_desired
        
        self._CMIP_Data_Collection(self.run_parallel)
        
    
    def _CMIP_Data_Collection(self,run_parallel):
        """
        obj._CMIP_Data_Collection
        
        Downloads all necessary CMIP6 files, performs location and temperature
        computation, collects temperature change values.
        """
        if self.world_map:
            self._World_map_plotting()
        
        if self.data_folder is None:
            folder_path = os.path.join(os.path.dirname(__file__),"..","..","..","CMIP6_Data_Files")
        else:
            folder_path = self.data_folder
        
        if not os.path.isabs(self.model_guide):
            self.model_guide = os.path.abspath(os.path.join(os.path.dirname(__file__),
                                                            self.model_guide))
        
        model_guide_wb = openpyxl.load_workbook(self.model_guide)
        #Downloads necessary files for each scenario
        for scenario in self.scenario_list:
            # we want for this large amount of data to be outside the MEWS
            # code base. We also want for it to stay in place once it is 
            # downloaded so that the user will not have to repeatedly wait for 
            # the large amount of downloading to occur.
             
            model_guide_ws = model_guide_wb[scenario]
            
            #Creates folders
            if os.path.exists(os.path.join(folder_path)) == False:
                os.mkdir(os.path.join(folder_path))
            if os.path.exists(os.path.join(folder_path,scenario)) == False:
                os.mkdir(os.path.join(folder_path,scenario))
        
            if run_parallel:
                #Performs all file downloads with threading
                threads = [] 
            
            # Bound the number of threads that can execute at once.
            semaphore = thr.Semaphore(self._num_cpu)
            
            # This process tends to hang and you have to check it 
            # and see if it is working 
            for row in range(2,model_guide_ws.max_row+1):
                model_name = model_guide_ws.cell(row,1).value
                
                # do not download if a gcm is not included
                if model_name in self.gcm_to_skip:
                    continue
                
                link = model_guide_ws.cell(row,2).value
                
                file_path_list = link.split("/")
                file_name = file_path_list[-1]
                path = os.path.join(folder_path,scenario,file_name)
                
                # only download (takes awhile) if the path does not exist. Otherwise, leave the files in place!
                if not os.path.exists(path):                
                    if run_parallel:  
                        try:
                            t = thr.Thread(target=self._Download_File,args=[path, link, semaphore])
                            t.start()
                            threads.append(t)
                        except:
                            logger.info("Parallel processing failed for threading! Reverting to non-parallel")
                            #start over
                            self._CMIP_Data_Collection(False)
                    else:
                        self._Download_File(path,link,None)
                    
            if run_parallel:
                for thread in threads:
                    thread.join()
                    
            logger.info(f"All {scenario} files downloaded")
                    
        #Runs bulk of calculations with multiprocessing
        if run_parallel: 
            pool = mp.Pool(self._num_cpu)
            total_model_list = []
        
        self.total_model_data = {}
        for idx,scenario in enumerate(self.scenario_list):
            if run_parallel:
                total_model_list.append(pool.apply_async(self._Compiling_Data,(model_guide_wb,folder_path,scenario,)))
            else:
                self.total_model_data[scenario] = self._Compiling_Data(model_guide_wb,folder_path,scenario)
        
        if run_parallel:
            pool.close()
            pool.join()
            for index in range(len(self.scenario_list)):
                self.total_model_data[total_model_list[index].get().scenario] = total_model_list[index].get()
          
        model_guide_wb.close()
        logger.info("Temperature computation complete")
        
        #calculates all result statistics used for delT and results1()
        self._calculate_stats()
        
        #Creates dictionary of temperature change results
        self.delT = {scenario: self.total_model_data[scenario].delT_list_reg[self.end_year-self.baseline_year] for scenario in self.scenario_list if scenario != "historical"}


    def _World_map_plotting(self):
        """
        obj._World_map_plotting()
        
        Plots selected latitude-longitude point on a world map.
        """
        if self.use_cartopy and self.display_plots:
            ##### Potting Location for Reference #####
            plt.figure(1, figsize=[30,13])
            ax = plt.axes(projection=ccrs.PlateCarree())
            ax.stock_img()
            plt.plot([self.lon_desired,self.lon_desired],
                     [self.lat_desired,self.lat_desired-0.5],
                     color='red', linestyle='-',linewidth=10, 
                     transform=ccrs.PlateCarree())
            ax.add_feature(cf.BORDERS)
            ax.coastlines()
        else:
            logger.warning("Cannot display world map because cartopy is not properly installed")
        
        
    def _Download_File(self,path,link,semaphore):
        """
        obj._Download_File(path,link)
        
        Downloads a file from CMIP6. If errors are encountered, other data 
        nodes will be checked for a duplicate file.
        """
        def loc_download_func(path,link):
            if os.path.exists(path) == False:
                logger.info("Downloading file from:",link)
                logger.info("Downloading file to:",path)
                try: 
                    urllib.request.urlretrieve(link, path)
                except:
                    logger.warning("Download failed. Data node is down. Trying again with other nodes")
                    all_data_nodes = ["aims3.llnl.gov",
                                      "esgf-data.ucar.edu",
                                      "esgf-data1.llnl.gov",
                                      "dpesgf03.nccs.nasa.gov"]
                    # international download locations we will avoid.
                                      # "esgf-data2.diasjp.net",
                                      # "esgf-data3.diasjp.net",
                                      # "esgf-node2.cmcc.it",
                                      # "vesg.ipsl.upmc.fr",
                                      # "cmip.dess.tsinghua.edu.cn"
                                      # "cmip.bcc.cma.cn",
                                      # "crd-esgf-drc.ec.gc.ca",
                                      # "dist.nmlab.snu.ac.kr",
                                      # "esg.lasg.ac.cn",
                                      # "esg-cccr.tropmet.res.in",
                                      # "esg-dn1.nsc.liu.se",
                                      # "esg-dn2.nsc.liu.se",
                                      # "esgf.nci.org.au",
                                      # "esgf3.dkrz.de",
                    for node in all_data_nodes:
                        try:
                            link = link.lstrip("http://")
                            link_list = link.split("/")
                            link_list[0] = node
                            link = "/".join(link_list)
                            link = "http://" + link
                            urllib.request.urlretrieve(link,path)
                            break
                        except:
                            pass
                    if node == all_data_nodes[-1]:
                        raise ConnectionError(f"Could not download file: {path}")  
        # Sempaphore bounds the number of tasks occuring at once.
        if semaphore is None:
            loc_download_func(path,link)
        else:
            with semaphore:
                loc_download_func(path,link)
                    
                    
    def _Compiling_Data(self,model_guide_wb,folder_path,scenario):
        """
        obj._Compiling_Data(model_guide_wb,folder_path,scenario)
        
        Compiles information from all files of a model and stores them as model
        objects. Then uses Total objects to store all information on the models
        for a scenario once calculations have been run.
        """
        scenario_model_data = {}
        model_guide_ws = model_guide_wb[scenario]
        data_list = []
        for row in range(2,model_guide_ws.max_row+1):
            model_name = model_guide_ws.cell(row,1).value
            if not model_name in self.gcm_to_skip:
                link = model_guide_ws.cell(row,2).value
                file_path_list = link.split("/")
                file_name = file_path_list[-1] 
                path = os.path.join(folder_path,scenario,file_name)
                file_count = model_guide_ws.cell(row,3).value
                file_num = model_guide_ws.cell(row,4).value
                Model = self.Model_object(model_name,path)
                if file_count == file_num:
                    if data_list != []:
                        scenario_model_data[Model.name] = Model
                        data_list.append(Model.data)
                        setattr(scenario_model_data[Model.name],'data', xr.concat(data_list,dim='time'))
                        data_list = []
                    else:
                        scenario_model_data[Model.name] = Model
                else:
                    data_list.append(Model.data)
        
        #Combines information into Total objects
        Total_obj = self.Total_data(scenario,self.baseline_year,self.end_year,self._cmip6_start_year,self._historical_scen_str)
        for model in scenario_model_data:
            scenario_model_data[model].set_data()
            self._yearly_dataset_temps(scenario_model_data[model],Total_obj)
        if self.calculate_error:
            Total_obj.compute_error()
        return Total_obj


    class Model_object:
        """
        Model_object class used for storing information from each model in objects.
        """
        def __init__(self,model,path):
            self.name = model
            dset = xr.open_dataset(path, decode_times=True, use_cftime=True) 
            self.data = dset
            #To be set later
            self.error = 0
            self.resolution = 0
        def set_data(self):
            start_year = str(self.data.coords['time'].values[0])
            start_year_list = start_year.split("-")
            self.start_year = int(start_year_list[0])
            end_year = str(self.data.coords['time'].values[-1])
            end_year_list = end_year.split("-")
            self.end_year = int(end_year_list[0])
            self.data = self.data['tas']
            
                    
    class Total_data:
        """
        Total_data class used for storing resulting data from all models.
        """
        def __init__(self,scenario,baseline_year,endyear,cmip_base_year,hist_str):
            if scenario != hist_str:
                self.start_year = baseline_year
                self.end_year = endyear
                self.year_temp_dict = {new_list: [] for new_list in range(baseline_year,endyear+1,1)}
            else:
                self.start_year = cmip_base_year
                self.end_year = baseline_year
                self.year_temp_dict = {new_list: [] for new_list in range(cmip_base_year,baseline_year+1,1)}
            
            self.scenario = scenario
            self.avg_error_list = []
            num_years = self.end_year - self.start_year + 1
            self.delT_list = list(np.zeros(num_years))
            self.avg_list = np.zeros(num_years)
            self.CI_list = np.zeros(num_years)
            self.delT_list_reg = None
            self.delT_polyfit = None
            self.baseline_regression = None
            self.dataset = {}  # different lengths of data per year may be possible
            self.R2 = None
            
        def compute_error(self):
            self.avg_error = np.mean(self.avg_error_list)
            
            
    def _yearly_dataset_temps(self,model_object,total_object):
        """
        obj._yearly_dataset_temps(model_object,total_object)
        
        Calculates average yearly temperature value for the specific location
        input using 2-point interpolation and error calcualtions. Stores output
        data in the total scenario object.
        """
        data_tas = model_object.data
        
        start_year = model_object.start_year
        
        end_year = self.end_year
        if end_year > model_object.end_year:
            end_year = model_object.end_year
            #raise ValueError("The GCM model has an end year of"
            #                 +" {0:d} but an end year greater than this of {1:d} was requested".format(
            #                     model_object.end_year,end_year))            
        
        local_max_diff = 0
        total_max_diff = 0
        years = np.linspace(start_year,end_year,num=end_year-start_year+1)
        years = [int(item) for item in years]
        for year_index in range(end_year-start_year+1):
            mon_temp = [0]*12
            for month in range(1,13):
                data = data_tas.isel(time=int((years[year_index]-start_year)*12+month-1))
                lon_data_coords = data.coords['lon'].values
                lat_data_coords = data.coords['lat'].values
                
                #Run for first month of the first year because location does not change
                if month == 1 and year_index == 0:
                    
                    #Checks if the latitude input is in range of the Global Climate Model grid
                    if self.lat_desired <= lat_data_coords[0] or self.lat_desired >= lat_data_coords[-1]:
                        raise ValueError("Latitude input is out of range of GCM")
                    
                    #Setting Arbitrary grid points if lon<358.5
                    lon_grid_coords_1 = lon_data_coords[-1]
                    lon_grid_coords_2 = lon_data_coords[0]
                    
                    #Finding the coords of the closest grid points
                    lon_init = lon_data_coords[0]
                    delta_lon = lon_data_coords[1]-lon_init
                    x = math.floor((self.lon_desired - lon_init)/delta_lon)
                
                    lat_init = lat_data_coords[0]
                    delta_lat = (lat_data_coords[-1]-lat_init)/(len(lat_data_coords)-1)
                    y = math.floor((self.lat_desired - lat_init)/delta_lat)
                    
                #Finding distances from the grid points
                if x >= len(lon_data_coords)-1: #Accounting for Longitude Wrapping
                    x = -1
                 
                lon_grid_coords_1 = lon_data_coords[x]
                lon_grid_coords_2 = lon_data_coords[x+1]
                
                if x == -1 and self.lon_desired >= 300: #For data points that are in between [-1] and [0] on the left of lon = 0
                    diff_lon = (lon_grid_coords_2 - (lon_grid_coords_1 - 360))
                    grid_lon_mult = abs(self.lon_desired-lon_grid_coords_1)/diff_lon
                elif x == -1: #For data points that are in between [-1] and [0] on the right of lon = 0
                    diff_lon = (lon_grid_coords_2 - (lon_grid_coords_1 - 360))
                    grid_lon_mult = (1-abs(lon_grid_coords_2-self.lon_desired)/diff_lon)
                else:
                    grid_lon_mult = abs(self.lon_desired-lon_grid_coords_1)/delta_lon
                
                lat_grid_coords_1 = lat_data_coords[y]
                grid_lat_mult = abs(self.lat_desired-lat_grid_coords_1)/delta_lat
                
                if grid_lat_mult > 1:
                    delta_lat = (lat_data_coords[-1]-lat_init)/(len(lat_data_coords)+1)
                    y = math.floor((self.lat_desired - lat_init)/delta_lat)
                    lat_grid_coords_1 = lat_data_coords[y]
                    grid_lat_mult = abs(self.lat_desired-lat_grid_coords_1)/delta_lat
                    
                #Location Error
                if month == 1 and year_index == 0 and self.calculate_error:

                    dist = np.sqrt(0.5) - np.sqrt((0.5-grid_lon_mult)**2 + (0.5-grid_lat_mult)**2)
                    dist_error = (1/np.sqrt(0.5))*dist
                    
                    #Normalization Error Boundaries.Set for the US
                    US_1_lon, US_1_lat = 235, 25
                    US_1_x,US_1_y = math.floor((US_1_lon - lon_init)/delta_lon), math.floor((US_1_lat - lat_init)/delta_lat)
    
                    US_4_lon, US_4_lat = 295, 50
                    US_4_x,US_4_y = math.ceil((US_4_lon - lon_init)/delta_lon), math.ceil((US_4_lat - lat_init)/delta_lat)
                    
                #Establishing temperature values for the 4 surrounding grid points
                temp_1 = float(data.isel(lon=x).isel(lat=y).values)
                temp_2 = float(data.isel(lon=x+1).isel(lat=y).values)
                temp_3 = float(data.isel(lon=x).isel(lat=y+1).values)
                temp_4 = float(data.isel(lon=x+1).isel(lat=y+1).values)     
                
                #Normalization Error
                if self.calculate_error:
                    US_temp_array = data.values[US_1_y:US_4_y,US_1_x:US_4_x] #[rows,cols] -> [y,x]   
                    if np.max(abs(np.diff(US_temp_array))) > total_max_diff:
                        total_max_diff = np.max(abs(np.diff(US_temp_array))) 
                    US_temp_array = np.rot90(US_temp_array)
                    if np.max(abs(np.diff(US_temp_array))) > total_max_diff:
                        total_max_diff = np.max(abs(np.diff(US_temp_array)))
                        
                    #local Normalization Error
                    local_temp_diff = np.max(abs(np.array([temp_1-temp_2,temp_1-temp_3,temp_3-temp_4,temp_2-temp_4])))
                    if local_temp_diff > local_max_diff:
                        local_max_diff = local_temp_diff
                
                temp_inter_1 = temp_1*(1-grid_lon_mult) + temp_2*grid_lon_mult
                temp_inter_2 = temp_3*(1-grid_lon_mult) + temp_4*grid_lon_mult
                temp_des = temp_inter_1*(1-grid_lat_mult) + temp_inter_2*grid_lat_mult
                
                mon_temp[month-1] = temp_des
            
            yearly_avg_temp = sum(mon_temp)/12
            #Skipping years past 2014 for some historical data sets for consistency
            if years[0] == self._cmip6_start_year and years[year_index] > self.baseline_year: continue

            total_object.year_temp_dict[years[year_index]].append(yearly_avg_temp)

        
        if self.calculate_error:
            norm_factor =local_max_diff/total_max_diff
            error = dist_error*norm_factor
        else:
            error = None
        model_object.error = error
        total_object.avg_error_list.append(error)
        
    
    def _regress_scenario(self,scenario, years,sigma0, func, p0, baseline, delT_at_baseline):
        """
        This function creates the regressions to CMIP6 data.
        
        """
        
        # y data in terms of absolute (abs) temperature
        y_abs_data_1d = np.concatenate([self.total_model_data[scenario].year_temp_dict[year] for year in years])
        num_points = [len(self.total_model_data[scenario].year_temp_dict[year]) for year in years]

        # calculate to delT from baseline year (i.e. 2014 for CMIP6) temperature.
        y_zero_data_1d = y_abs_data_1d - baseline

            
        # assemble a repeated entry of the years and sigma values.
        x_data_1d = np.concatenate([np.ones(num_point)*year for num_point, year in zip(num_points,years)])
        sigma = np.concatenate([np.ones(num_point)*sigma0[year-years[0]] for num_point,year in zip(num_points,years)])

        # the fit is performed to data that is 0 at 2014 mean of CMIP6 data.
        # we can then add 

        fit, _ = curve_fit(func, x_data_1d, y_zero_data_1d, p0, sigma=sigma)

        regression = func(np.array(years),*fit)
        regr_1d = np.concatenate([np.ones(num_point)*regression[year-years[0]] for num_point, year in zip(num_points,years)])
        
        # calculate R2 value for each fit.
        residuals = y_zero_data_1d - regr_1d
        ss_res = np.sum(residuals**2)
        mean_y = np.mean(y_zero_data_1d)
        ss_tot = np.sum((y_zero_data_1d - mean_y)**2)
        r_squared = 1 - (ss_res/ss_tot)
        
        
        self.total_model_data[scenario].delT_list_reg = regression
        self.total_model_data[scenario].delT_polyfit = fit
        self.total_model_data[scenario].R2 = r_squared
        self.total_model_data[scenario].y_data_1d = y_zero_data_1d # arranged for scatter plots
        self.total_model_data[scenario].x_data_1d = x_data_1d # arranged for scatter plots of the entire dataset.
        self.total_model_data[scenario].delT_data_1d = y_zero_data_1d + delT_at_baseline
        
        # calculate confidence intervals for each set of data per year. 
        for year_index in range(len(years)):
            dataset = self.total_model_data[scenario].year_temp_dict[years[year_index]] - baseline
            mean_dataset = np.mean(dataset)
            delT = dataset + delT_at_baseline
            lower_proj, upper_proj = st.t.interval(confidence=0.95, df=len(dataset)-1, loc=np.mean(dataset), scale=st.sem(dataset))
            CI = upper_proj - mean_dataset
            self.total_model_data[scenario].CI_list[year_index] = CI
            self.total_model_data[scenario].avg_list[year_index] = mean_dataset
            self.total_model_data[scenario].delT_list[year_index] = delT
            self.total_model_data[scenario].dataset[year_index] = dataset
    
    
    def _calculate_stats(self):
        """
        obj._calculate_stats()
        
        Calculates ensemble means and 95% confidence interval bounds for the 
        total temperature data of each scenario, then performs a fixed point
        regression on all data to produce useable results. The regressed data
        is later outputted as results.
        """
        
        baseline_year = self.baseline_year

        """
        
        Transition from absolute temperatures to delta T from the pre-industrial baseline
        from CMIP6 historic post-casts. MEWS regressions consider 0 to be the delT for the
        CMIP 6 baseline year (2014). It then adds the delT so that everything is
        directly in terms of 1850-1900 delT consistent with CMIP6
        
        """
        historical_years = [year for year in range(self._cmip6_start_year,baseline_year+1,1)]
        # we include 2014 in future years so that the zero target is enforced.
        future_years = [year for year in range(baseline_year,self.end_year+1,1)] 
        baseline_year_avg_temp = np.mean(self.total_model_data[self._historical_scen_str].year_temp_dict[baseline_year])
        #baseline-1 average
        baseline_year_m1_avg_temp = np.mean(self.total_model_data[self._historical_scen_str].year_temp_dict[baseline_year-1])
        
        hist_slope = baseline_year_avg_temp - baseline_year_m1_avg_temp
        
        
        preindustrial_baseline_temp = np.mean(np.concatenate([self.total_model_data[
            self._historical_scen_str].year_temp_dict[year] for year in 
            np.arange(self._cmip6_start_year,self._cmip6_baseline_end_year+1)]))
        self.preindustrial_baseline_temp = preindustrial_baseline_temp
        
        # this is the delT that must be added onto every regression to
        # actually get 1850 to 1900 delT numbers.
        baseline_year_delT = baseline_year_avg_temp - preindustrial_baseline_temp
        self.baseline_year_delT = baseline_year_delT
        
        # this function cannot be changed unless the "regression" function
        # is also changed in _regress_scenario
        def poly_function(x,*p):
            # makes value 0 for baseline year.
            return np.poly1d(p)(x-baseline_year)
        
        
        for scenario in self.scenario_list:
            
            # the number of initial values controls the order of the polynomial.
            
            initial_values = tuple(np.zeros(self._polynomial_fit_order[scenario]))
            
            if scenario == self._historical_scen_str:
                years = historical_years
                sigma0=np.ones(len(years))
                sigma0[-1] = 0.001 # fix the last point to zero target
            else:
                scen_avg = np.mean(self.total_model_data[scenario].year_temp_dict[self.baseline_year+1])
                # we give + 10 years so that the net trend is observed rather than local variations
                scen_avg_p1 = np.mean(self.total_model_data[scenario].year_temp_dict[self.baseline_year+10])
                scen_slope = (scen_avg_p1 - scen_avg)/9
                
                years = future_years
                sigma0=np.ones(len(years))
                if self._align_gcm_to_historical:
                    sigma0[0] = 0.001 # fix the first point to zero target
                    
                    # big assumption to give a smooth result!, Re-baseline to the 2014 historical average as
                    # the start for all scenarios - IS Shifting the data like this
                    # justified? The model groups for different scenarios are different
                    # but the discontinuities are large. IPCC shows a historic
                    # smooth relationship for global results.

                    avg_slope = (scen_slope + hist_slope)/2
                    target_temp = baseline_year_avg_temp + avg_slope
                    offset_fact = target_temp - scen_avg
                    for key,val in self.total_model_data[scenario].year_temp_dict.items():
                        self.total_model_data[scenario].year_temp_dict[key] = val + offset_fact
                        
                    
                    # add 2014 data to the future scenario regressions
                    self.total_model_data[scenario].year_temp_dict[
                        self.baseline_year] = self.total_model_data[
                            self._historical_scen_str].year_temp_dict[self.baseline_year]
                            
                else:
                    # add 2014 as a continuation of the slope for the first two years
                    self.total_model_data[scenario].year_temp_dict[
                        self.baseline_year] = self.total_model_data[
                            scenario].year_temp_dict[self.baseline_year+1]-scen_slope                    
                
            self._regress_scenario(scenario, 
                                   years, 
                                   sigma0, 
                                   poly_function, 
                                   initial_values, 
                                   baseline_year_avg_temp,
                                   baseline_year_delT)

                

    def results1(self,scatter_display=[False,False,False,False,False,False],
                 regression_display=[True,True,True,True,True,True],
                 CI_display=[True,True,True,True,True,True],
                 plot_begin_year=1950,write_png=None,show_plot=False):
        """
        Plots the temperature change values for the scenarios for the years 1950-2100.

        >>> obj.results1(scatter_display=[False,False,False,False,False,False],
                         regression_display=[True,True,True,True,True,True],
                         CI_display=[True,False,True,False,True,False])
        
        Parameters
        ----------
        
        scenario_list : str : list : optional : Default = 
        
        scatter_display : bool : list : optional : Default = [False,False,False,False,False,False]
            A list of 6 boolean values used for selecting which scatter data to plot.
            The list items correspond to: historical, SSP119, SSP126, SSP245, SSP370, SSP585.
            
        regression_display : bool : list : optional : Default = [True,True,True,True,True,True]
            A list of 6 boolean values used for selecting which regressions to plot.
            The list items correspond to: historical, SSP119, SSP126, SSP245, SSP370, SSP585.
            
        CI_display : bool : list : optional : Default = [True,False,True,False,True,False]
            A list of 6 boolean values used for selecting which Confidence intervals to plot.
            The list items correspond to: historical, SSP119, SSP126, SSP245, SSP370, SSP585.       
            
        plot_begin_year : int : optional : Default = 1950
            Control what date the plot left hand side begins at.
            
        write_png : str : optional : default = None
            if not None, write a .png file to the name/path location indicated 
            by write_png
            
        show_plot : bool : optional : default = False
            if True, shows the plot being made rather than only creating the 
            plot object
        
        Returns
        -------
        None

        """
        
        if self.display_plots: 
            historical_years = [year for year in range(self._cmip6_start_year,self.baseline_year+1,1)]
            future_years = [year for year in range(self.baseline_year,self.end_year+1,1)]
            
            plt.rcParams.update({'font.size': 22})
            fig,ax = plt.subplots()
            
            
            #TODO - generalize this to be part of _valid_scenarios names.
            #years, colors, and labels organized into dictionaries for succinctness
            year_dict = {self._historical_scen_str:historical_years,"SSP119":future_years,"SSP126":future_years,"SSP245":future_years,"SSP370":future_years,"SSP585":future_years}
            color_dict = {self._historical_scen_str:"k","SSP119":"dodgerblue","SSP126":"navy","SSP245":"gold","SSP370":"red","SSP585":"maroon"}
            scatter_label_dict = {self._historical_scen_str:"Averaged Historical Data","SSP119":"Averaged SSP1-1.9 Data","SSP126":"Averaged SSP1-2.6 Data",
                                  "SSP245":"Averaged SSP2-4.5 Data","SSP370":"Averaged SSP3.70 Data","SSP585":"Averaged SSP5-8.5 Data"}
            regression_label_dict = {self._historical_scen_str:"Historical Regression","SSP119":"SSP1-1.9 Regression","SSP126":"SSP1-2.6 Regression",
                                     "SSP245":"SSP2-4.5 Regression","SSP370":"SSP3.70 Regression","SSP585":"SSP5-8.5 Regression"}
            CI_label_dict = {self._historical_scen_str:"Historical Data 95% CI","SSP119":"SSP1-1.9 Data 95% CI","SSP126":"SSP1-2.6 Data 95% CI",
                             "SSP245":"SSP2-4.5 Data 95% CI","SSP370":"SSP3.70 Data 95% CI","SSP585":"SSP5-8.5 Data 95% CI"}
    
            flat_df_list = []
            for scenario in self.scenario_list:
                
                index = self.scenario_list.index(scenario)
                
                if scatter_display[index]:

                    ax.scatter(self.total_model_data[scenario].x_data_1d,
                               self.total_model_data[scenario].y_data_1d,
                               c=color_dict[scenario],
                               label=scatter_label_dict[scenario])
                if regression_display[index]:
                    ax.plot(year_dict[scenario],
                            self.total_model_data[scenario].delT_list_reg,
                            color_dict[scenario],linewidth=3,
                            label=regression_label_dict[scenario])
                if CI_display[index]:
                    ax.fill_between(year_dict[scenario],
                                    (self.total_model_data[scenario].avg_list-self.total_model_data[scenario].CI_list),
                                    (self.total_model_data[scenario].avg_list+self.total_model_data[scenario].CI_list),
                                    color=color_dict[scenario],alpha=.2,label=CI_label_dict[scenario])
            
                [flat_df_list.append([yr, val, "gcm data point", scenario])  
                 for yr,val in zip(self.total_model_data[scenario].x_data_1d, 
                                   self.total_model_data[scenario].y_data_1d)]
                
                [flat_df_list.append([yr, val, "polynomail fit", scenario])  
                 for yr,val in zip(year_dict[scenario], 
                                   self.total_model_data[scenario].delT_list_reg)]
                
                [flat_df_list.append([yr, val, "1/2 symmetric 95% confidence interval (CI) (add and subtract this from average to get CI)", scenario])  
                 for yr,val in zip(year_dict[scenario], 
                                   self.total_model_data[scenario].CI_list)]
                                   
                [flat_df_list.append([yr, val, "average of gcm data points", scenario])  
                 for yr,val in zip(year_dict[scenario], 
                                   self.total_model_data[scenario].avg_list)]
                
            df = pd.DataFrame(flat_df_list, columns=["Year","Change in Temperature (°C)","Type","Scenario"])
            
            df.to_csv(write_png + ".csv")
                
            plot_box = ax.get_position()
            ax.set_position([plot_box.x0, plot_box.y0,plot_box.width*0.8,plot_box.height])
            ax.legend(loc='center left', bbox_to_anchor=(1, 0.5))
            plt.xlabel("Year",fontdict={'fontsize': 26})
            plt.ylabel("Change in Temperature from {0:d} (°C)".format(self.baseline_year),fontdict={'fontsize': 26})
            plt.title("Surface Temperature Change at lat=({0:.4f}, lon={1:.4f})".format(self.lat_desired,self.lon_desired,self.baseline_year),y=1.02,x=0.7,fontdict={'fontsize': 30})
            plt.xlim(plot_begin_year,self.end_year)
            plt.grid()
            fig.set_size_inches(18.5, 10.5)
            

            if not write_png is None:
                fig.savefig(write_png,dpi=300)
            if show_plot:  
                plt.show()  
                
            return fig,ax
    
    def _total_model_data_to_csv(self,year_dict):
        pass
    
    def results2(self,desired_scenario,resolution='low'):
        """
        Plots the temperature change heat map of the United States from baseline_year
        to end_year for a selected scenario.
        
        >>> obj.results2(desired_scenario,
                         resolution='low')
        
        Parameters
        ----------
        
        desired_scenario : str 
            A str that selectes the future scenario heat map that will be calculated and
            displayed. The possible strings are: "SSP119", "SSP126", "SSP245", "SSP370", "SSP585".
            Any other strings will return an error.
            
        resolution : str : optional : Default = 'low'
            A string that selectes the resolution of the heat map plotted. The default 
            setting is low so the simulation can be run quicker. Higher qualities can
            take much longer to calculate and display. All possible resolutions are:
            "low", "medium", "high", "extreme". "test" is very inaccurate; only used for
            unittesting.
            
        Returns
        -------
        None
        
        """
        resolution_dict = {"test":[2,2],"low":[60,27],"medium":[120,54],"high":[240,108],"extreme":[480,216]}
 
        lon_array = np.linspace(360-125,360-66,num=resolution_dict[resolution][0])
        lat_array = np.linspace(24,50,num=resolution_dict[resolution][1])
        folder_path = self.data_folder
        
        result_arrays=[]
        for scenario in ['historical',desired_scenario]:
            year = self.baseline_year if scenario == 'historical' else self.end_year
    
            
            model_guide_wb = openpyxl.load_workbook(os.path.join(self.model_guide))
            model_guide_ws = model_guide_wb[scenario]
        
            model_list = []
            for row in range(2,model_guide_ws.max_row+1):
                name = model_guide_ws.cell(row,1).value
                if (name not in model_list) and (name not in self.gcm_to_skip) :
                    model_list.append(name)
            
            if self.run_parallel:
                pool = mp.Pool(self._num_cpu)
            temp_array_list = []
            for model in model_list:
                if self.run_parallel:
                    temp_array_list.append(pool.apply_async(self._results2_calc,
                                                        (model_guide_ws,
                                                         folder_path,
                                                         scenario,
                                                         model,
                                                         lat_array,
                                                         lon_array,
                                                         year)))
                else:
                    temp_array_list.append(self._results2_calc(model_guide_ws,
                                        folder_path,
                                        scenario,
                                        model,
                                        lat_array,
                                        lon_array,
                                        year))
            
            if self.run_parallel:
                pool.close()
                pool.join()
        
                for index in range(len(temp_array_list)):
                    temp_array_list[index] = temp_array_list[index].get()
        
            array = temp_array_list[0]
            if len(model_list) > 1:
                for index in range(1,len(model_list)):
                    array  += temp_array_list[index]
                array = array/len(model_list)
            result_arrays.append(array)
        
        baseline = xr.DataArray(data=result_arrays[0],dims=["lat", "lon"],coords=[lat_array,lon_array])
        projection = xr.DataArray(data=result_arrays[1],dims=["lat", "lon"],coords=[lat_array,lon_array])
        delta_temp = projection - baseline

        value_high = np.max(delta_temp)
        value_low = abs(np.min(delta_temp))
        limit = np.max([math.ceil(value_high),math.ceil(value_low)])

        extent = [-125, -66, 24, 46]
        plt.rcParams.update({'font.size': 12})
        if self.display_plots:
            if self.use_cartopy:
                plt.figure(figsize=(16, 6))
                ax = plt.subplot(1, 1, 1, projection=ccrs.PlateCarree())
                delta_temp.plot.pcolormesh(ax=ax,cmap='coolwarm',vmin=-limit,vmax=limit)
                ax.gridlines(draw_labels=True)
                ax.set_extent(extent)
                ax.coastlines()
                ax.add_feature(cf.BORDERS)
                ax.add_feature(cf.STATES)
                plt.title(f"US Temperature Change from {self.baseline_year} to {self.end_year} According to {desired_scenario}",x=0.55,fontdict={'fontsize': 16})
            else:
                logger.warning("Cannot display US borders because cartopy is not properly installed")
                fig = plt.figure()
                ax = plt.axes([0,0,2,1])
                c = ax.pcolormesh(lon_array, lat_array, delta_temp, cmap='coolwarm', vmin=-limit, vmax=limit)
                ax.set_title('pcolormesh')
                fig.colorbar(c, ax=ax)
                plt.ylabel("Latitude [Degrees North]")
                plt.xlabel("Longitude [Degrees East]")
                plt.xlim(235,294)
                plt.ylim(24,50)
                plt.grid()
                plt.title(f"US Temperature Change from {self.baseline_year} to {self.end_year} According to {desired_scenario}")
                
            plt.show()

            
    def _results2_calc(self,model_guide_ws,folder_path,scenario,model,lat_array,lon_array,year):
        """
        obj._results2_calc(model_guide_ws,
                           folder_path,
                           scenario,
                           model,
                           lat_array,
                           lon_array,
                           year)
        
        Performs all of the calculations associated with plotting the US heat maps
        in the function call obj.results2(desired_scenario, resolution). First, it
        compiles the data used, then computes the temperature change grid array, 
        and finally returns that data to results2.
        """
        data_list = []
        for row in range(2,model_guide_ws.max_row+1):
            model_name = model_guide_ws.cell(row,1).value
            link = model_guide_ws.cell(row,2).value
            file_path_list = link.split("/")
            file_name = file_path_list[-1] 
            path = os.path.join(folder_path,scenario,file_name)
            if model_name == model:
                Model = self.Model_object(model_name,path)
                file_count = model_guide_ws.cell(row,3).value
                file_num = model_guide_ws.cell(row,4).value
                if file_count == file_num:
                    if data_list != []:
                        Complete_model = Model
                        data_list.append(Model.data)
                        setattr(Complete_model,'data', xr.concat(data_list,dim='time'))
                    else:
                        Complete_model = Model
                else:
                    data_list.append(Model.data)
        
        Complete_model.set_data()
        temperature_baseline = np.zeros((len(lat_array),len(lon_array)))
        for i in range(len(lat_array)):
            for j in range(len(lon_array)):
                lat = lat_array[i]
                lon = lon_array[j]
            
                mon_temp = [0]*12
                for month in range(1,13):
                    data = Complete_model.data.isel(time=int((year-Complete_model.start_year)*12+month-1))
                    lon_data_coords = data.coords['lon'].values
                    lat_data_coords = data.coords['lat'].values

                    if month == 1:
                        lon_init = lon_data_coords[0]
                        delta_lon = lon_data_coords[1]-lon_init
                        x = math.floor((lon - lon_init)/delta_lon)
                    
                        lat_init = lat_data_coords[0]
                        delta_lat = (lat_data_coords[-1]-lat_init)/(len(lat_data_coords)-1)
                        y = math.floor((lat - lat_init)/delta_lat)
   
                    lon_grid_coords_1 = lon_data_coords[x]
                    grid_lon_mult = abs(lon-lon_grid_coords_1)/delta_lon
                    lat_grid_coords_1 = lat_data_coords[y]
                    grid_lat_mult = abs(lat-lat_grid_coords_1)/delta_lat
                    
                    if grid_lat_mult > 1:
                        delta_lat = (lat_data_coords[-1]-lat_init)/(len(lat_data_coords)+1)
                        y = math.floor((lat - lat_init)/delta_lat)
                        lat_grid_coords_1 = lat_data_coords[y]
                        grid_lat_mult = abs(lat-lat_grid_coords_1)/delta_lat
                        
                    #Establishing temperature values for the 4 surrounding grid points
                    temp_1 = float(data.isel(lon=x).isel(lat=y).values)
                    temp_2 = float(data.isel(lon=x+1).isel(lat=y).values)
                    temp_3 = float(data.isel(lon=x).isel(lat=y+1).values)
                    temp_4 = float(data.isel(lon=x+1).isel(lat=y+1).values)     
                
                    temp_inter_1 = temp_1*(1-grid_lon_mult) + temp_2*grid_lon_mult
                    temp_inter_2 = temp_3*(1-grid_lon_mult) + temp_4*grid_lon_mult
                    temp_des = temp_inter_1*(1-grid_lat_mult) + temp_inter_2*grid_lat_mult
                    
                    mon_temp[month-1] = temp_des
                yearly_avg_temp = sum(mon_temp)/12
                temperature_baseline[i][j] = yearly_avg_temp
                
        return temperature_baseline
    

        

        

if __name__ == '__main__':
    pass



"""
Cities chosen based on climate zones
Miami lat: 25.7881 Lon: -80.31694
Houston lat: 29.9844 lon: -95.36072
Phooenix lat: 33.4278 lon: -112.0037
Atlanta lat: 33.6297 lon: -84.44224
Los Angeles lat: 33.9382 lon: -118.3866
Las Vegas lat: 36.2121 lon: -115.1939
San Francisco lat: 37.6196 lon: -122.3656
Baltimore lat: 39.1733 lon: -76.68408
Albuquerque lat: 35.0433 lon: -106.6129
Seattle lat: 47.4447 lon: -122.3144
Chicago lat: 41.7841 lon: -87.75514
Denver lat: 39.8466 lon: -104.6562
Minneapolis lat: 44.8852 lon: -93.23133
Helena lat: 46.6044 lon: -111.9892
Duluth lat: 46.8436 lon: -92.18658
Fairbanks lat: 64.8031 lon: -147.8761
"""
