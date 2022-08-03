# -*- coding: utf-8 -*-
"""

EPW - Lightweight Python package for editing EnergyPlus Weather (epw) files
EPW is not the EPW downloaded from pypi.org, it must be downloaded
from https://github.com/building-energy/epw

EPW License
===========

MIT License

Copyright (c) 2019 Building Energy Research Group (BERG)

Permission is hereby granted, free of charge, to any person obtaining a copy
of this software and associated documentation files (the "Software"), to deal
in the Software without restriction, including without limitation the rights
to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
copies of the Software, and to permit persons to whom the Software is
furnished to do so, subject to the following conditions:

The above copyright notice and this permission notice shall be included in all
copies or substantial portions of the Software.

THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
SOFTWARE.


Created on Wed Jun 20 13:00:31 2022

@author: tschoste
"""
###May need to be Changed ^^^^####

import openpyxl
import os
import warnings
import urllib.request
import time
import xarray as xr
import numpy as np
import math
import matplotlib.pyplot as plt
import scipy.stats as st
from scipy.optimize import curve_fit
import multiprocessing as mp
import threading as thr
import logging

########## Setting Proxies ##############
# Needed if downloading files 
# within Sandia
proxy = 'http://proxy.sandia.gov:80'
os.environ['http_proxy'] = proxy 
os.environ['HTTP_PROXY'] = proxy
os.environ['https_proxy'] = proxy
os.environ['HTTPS_PROXY'] = proxy


########## Initial Startup ##############
warnings.simplefilter(action='ignore') #Removes depreciation warnings


class CMIP_Data():
    """
    Calculates temperature change values from a baseline year to a future year
    for 5 SSP scenarios for a specific latitude-longitude input. Can also be 
    used to generate global warming plots. Used for CMIP6 data collection when
    accounting for global warming in MEWS.
    """
    def __init__(self, lat_desired, 
                 lon_desired,
                 year_baseline,
                 year_desired,
                 model_guide,
                 file_path,
                 data_folder=None,
                 scenario_list=["historical","SSP119","SSP126","SSP245","SSP370","SSP585"],
                 world_map=False,
                 calculate_error=True,
                 display_logging=False):
        
        """
        obj = CMIP_Data(lat_desired,
                 lon_desired,
                 year_baseline,
                 year_desired,
                 model_guide,
                 world_map,
                 calculate_error=True)
        
        Parameters
        ----------
        
        lat_desired : float
            Latitude value used for temperature change calculations. Must be
            between -90 and 90 [degrees].
        
        lon_desired : float
            Longitude value used for temperature change calculations. Must be 
            between -360 (when using degrees W) and 360 (when using degrees E).
            
        year_baseline : int
            Year input for the baseline temperature change. All delta_T's will 
            be calculated from this year. 1850 <= year <= 2014.
            
        year_desired : int
            Future year input for which all temperature change values will be 
            calculated to. 2015 <= year <= 2100.
            
        model_guide : str
            Filename for excel spreadsheet which contains the names and links 
            to all models used in the calculations.
            
        file_path : str
            Defines where the file is being called from.
            
        data_folder : str : optional : Default = None
            The name of the folder where the model_guide can be found and where
            all of the CMIP data files will downloaded to relative to the file
            the script is called from. If None, the code will search for the 
            model_guide and download all files to the cwd.
            
        world_map : bool : optional : Default = True
            Displays the latitude-longitude location desired on a world map.
            Can be useful in double checking the location is selected correctly.
            
        calculate_error : bool : optional : Default = True
            If true, will calculate normalization and distance error for all
            models. Sighlty faster to run if set to False.
            
        display_logging : bool : optional : Default = False
            If set to True, the logging statements will be printed to the console
            as well as saved to a file in the current directory.
            
        
        Returns
        -------
        
        None - Results found in obj.del_T which returns a dictionary of the 
        temperature change values for each scenario based on year_desired.
        
        Plots can be generated through obj.results1 and obj.results2
        """
        if data_folder != None: 
            self.dirname = os.path.join(os.path.abspath(os.getcwd()),data_folder)
        else:
            self.dirname = os.path.abspath(os.getcwd())
            
        self.world_map = world_map
        self.lat_desired = lat_desired
        self.lon_desired = lon_desired
        self.year_baseline = year_baseline
        self.year_desired = year_desired
        self.calculate_error = calculate_error
        self.model_guide = model_guide
        self.data_folder = data_folder
        self.scenario_list = scenario_list

        #Initializing logging
        global logger
        logger = logging.getLogger(__name__)
        logger.setLevel(logging.INFO)
        formatter = logging.Formatter("%(asctime)s | %(levelname)s | %(message)s")
        logger.handlers = []

        file_handler = logging.FileHandler("CMIP6_Data_Log.log")
        file_handler.setFormatter(formatter)
        logger.addHandler(file_handler)
        logger.propagate = False
        if display_logging:
            stream_handler = logging.StreamHandler()
            stream_handler.setFormatter(formatter)
            logger.addHandler(stream_handler)
            logger.propagate = False
        logger.info("Script Initiation")
        
        #Testing Cartopy module import
        try:
            global ccrs
            global cf
            import cartopy.crs as ccrs
            import cartopy.feature as cf
            self.use_cartopy = True
            urllib.request.urlretrieve("https://testingcaseinwhichthisfails")
        except:
            logger.warning("Unable to import cartopy")
            self.use_cartopy = False
        
        #Checking latitude and longitude inputs
        if self.lon_desired < 0: #If negative lon is inputted
            self.lon_desired += 360 
        if self.lon_desired > 360 or self.lon_desired < 0:
            raise ValueError("Longitude input is out of range")
        if self.lat_desired > 90 or self.lat_desired < -90:
            raise ValueError("Latitude input is out of range")
        
        self._CMIP_Data_Collection()
    
    def _CMIP_Data_Collection(self):
        """
        obj._CMIP_Data_Collection
        
        Downloads all necessary CMIP6 files, performs location and temperature
        computation, collects temperature change values.
        """
        if self.world_map:
            self._World_map_plotting()
        
        #Downloads necessary files for each scenario
        for scenario in self.scenario_list:
            folder_path = "CMIP6_Data_Files" 
            model_guide_wb = openpyxl.load_workbook(os.path.join(self.dirname,self.model_guide))
            model_guide_ws = model_guide_wb[scenario]
            
            #Creates folders
            if os.path.exists(os.path.join(self.dirname,folder_path)) == False:
                os.mkdir(os.path.join(self.dirname,folder_path))
            if os.path.exists(os.path.join(self.dirname,folder_path,scenario)) == False:
                os.mkdir(os.path.join(self.dirname,folder_path,scenario))
        
            #Performs all file downloads with threading
            threads = []    
            for row in range(2,model_guide_ws.max_row+1):
                link = model_guide_ws.cell(row,2).value
                file_path_list = link.split("/")
                file_name = file_path_list[-1] 
                path = os.path.join(self.dirname,folder_path,scenario,file_name)
                
                t = thr.Thread(target=self._Download_File,args=[path, link])
                t.start()
                threads.append(t)
            for thread in threads:
                thread.join()
            logger.info(f"All {scenario} files downloaded")
                    
        #Runs bulk of calculations with multiprocessing
        pool = mp.Pool(mp.cpu_count()-1)
        total_model_list = []
        for scenario in self.scenario_list:
            total_model_list.append(pool.apply_async(self._Compiling_Data,(model_guide_wb,folder_path,scenario,)))
        
        pool.close()
        pool.join()
        
        self.total_model_data = {}
        for index in range(len(self.scenario_list)):
            self.total_model_data[total_model_list[index].get().scenario] = total_model_list[index].get()
          
        model_guide_wb.close()
        logger.info("Temperature computation complete")
        
        #calculates all result statistics used for delT and results1()
        self._calculate_stats()
        
        #Creates dictionary of temperature change results
        self.delT = {scenario: self.total_model_data[scenario].delT_list_reg[self.year_desired-2015] for scenario in self.scenario_list if scenario != "historical"}


    def _World_map_plotting(self):
        """
        obj._World_map_plotting()
        
        Plots selected latitude-longitude point on a world map.
        """
        if self.use_cartopy:
            ##### Potting Location for Reference #####
            plt.figure(1, figsize=[30,13])
            ax = plt.axes(projection=ccrs.PlateCarree())
            ax.stock_img()
            plt.plot([self.lon_desired,self.lon_desired],
                     [self.lat_desired,self.lat_desired-0.5],
                     color='red', linestyle='-',linewidth=10, 
                     transform=ccrs.PlateCarree())
            ax.add_feature(cf.BORDERS)
            ax.coastlines()
            plt.show()
        else:
            logger.warning("Cannot display world map because cartopy is not properly installed")
        
        
    def _Download_File(self,path,link):
        """
        obj._Download_File(path,link)
        
        Downloads a file from CMIP6. If errors are encountered, other data 
        nodes will be checked for a duplicate file.
        """
        if os.path.exists(path) == False:
            logger.info("Downloading file:",path)
            try: 
                urllib.request.urlretrieve(link, path)
            except:
                logger.warning("Download failed. Data node is down. Trying again with other nodes")
                all_data_nodes = ["aims3.llnl.gov",
                                  "cmip.bcc.cma.cn",
                                  "crd-esgf-drc.ec.gc.ca",
                                  "dist.nmlab.snu.ac.kr",
                                  "esg.lasg.ac.cn",
                                  "esg-cccr.tropmet.res.in",
                                  "esg-dn1.nsc.liu.se",
                                  "esg-dn2.nsc.liu.se",
                                  "esgf.nci.org.au",
                                  "esgf3.dkrz.de",
                                  "esgf-data.ucar.edu",
                                  "esgf-data1.llnl.gov",
                                  "esgf-data2.diasjp.net",
                                  "esgf-data3.diasjp.net",
                                  "esgf-node2.cmcc.it",
                                  "vesg.ipsl.upmc.fr",
                                  "cmip.dess.tsinghua.edu.cn",
                                  "dpesgf03.nccs.nasa.gov"]
                for node in all_data_nodes:
                    try:
                        link = link.lstrip("http://")
                        link_list = link.split("/")
                        link_list[0] = node
                        link = "/".join(link_list)
                        link = "http://" + link
                        urllib.request.urlretrieve(link,path)
                        break
                    except:
                        pass
                if node == all_data_nodes[-1]:
                    raise ConnectionError(f"Could not download file: {path}")  
                    
                    
    def _Compiling_Data(self,model_guide_wb,folder_path,scenario):
        """
        obj._Compiling_Data(model_guide_wb,folder_path,scenario)
        
        Compiles information from all files of a model and stores them as model
        objects. Then uses Total objects to store all information on the models
        for a scenario once calculations have been run.
        """
        scenario_model_data = {}
        model_guide_ws = model_guide_wb[scenario]
        data_list = []
        for row in range(2,model_guide_ws.max_row+1):
            model_name = model_guide_ws.cell(row,1).value
            link = model_guide_ws.cell(row,2).value
            file_path_list = link.split("/")
            file_name = file_path_list[-1] 
            path = os.path.join(self.dirname,folder_path,scenario,file_name)
            file_count = model_guide_ws.cell(row,3).value
            file_num = model_guide_ws.cell(row,4).value
            Model = self.Model_object(model_name,path)
            if file_count == file_num:
                if data_list != []:
                    scenario_model_data[Model.name] = Model
                    data_list.append(Model.data)
                    setattr(scenario_model_data[Model.name],'data', xr.concat(data_list,dim='time'))
                    data_list = []
                else:
                    scenario_model_data[Model.name] = Model
            else:
                data_list.append(Model.data)
        
        #Combines information into Total objects
        Total_obj = self.Total_data(scenario)
        for model in scenario_model_data:
            scenario_model_data[model].set_data()
            self._yearly_dataset_temps(scenario_model_data[model],Total_obj)
        if self.calculate_error:
            Total_obj.compute_error()
        return Total_obj


    class Model_object:
        """
        Model_object class used for storing information from each model in objects.
        """
        def __init__(self,model,path):
            self.name = model
            dset = xr.open_dataset(path, decode_times=True, use_cftime=True) 
            self.data = dset
            #To be set later
            self.error = 0
            self.resolution = 0
        def set_data(self):
            start_year = str(self.data.coords['time'].values[0])
            start_year_list = start_year.split("-")
            self.start_year = int(start_year_list[0])
            end_year = str(self.data.coords['time'].values[-1])
            end_year_list = end_year.split("-")
            self.end_year = int(end_year_list[0])
            self.data = self.data['tas']
            
                    
    class Total_data:
        """
        Total_data class used for storing resulting data from all models.
        """
        def __init__(self,scenario):
            if scenario != "historical":
                self.start_year = 2015
                self.end_year = 2100
                self.year_temp_dict = {new_list: [] for new_list in range(2015,2101,1)}
            else:
                self.start_year = 1850
                self.end_year = 2014
                self.year_temp_dict = {new_list: [] for new_list in range(1850,2015,1)}
            
            self.scenario = scenario
            self.avg_error_list = []
            num_years = self.end_year - self.start_year + 1
            self.delT_list = np.zeros(num_years)
            self.CI_list = np.zeros(num_years)
            self.delT_list_reg = None
            
        def compute_error(self):
            self.avg_error = np.mean(self.avg_error_list)
            
            
    def _yearly_dataset_temps(self,model_object,total_object):
        """
        obj._yearly_dataset_temps(model_object,total_object)
        
        Calculates average yearly temperature value for the specific location
        input using 2-point interpolation and error calcualtions. Stores output
        data in the total scenario object.
        """
        data_tas = model_object.data
        start_year = model_object.start_year
        end_year = model_object.end_year
        local_max_diff = 0
        total_max_diff = 0
        years = np.linspace(start_year,end_year,num=end_year-start_year+1)
        years = [int(item) for item in years]
        for year_index in range(end_year-start_year+1):
            mon_temp = [0]*12
            for month in range(1,13):
                data = data_tas.isel(time=int((years[year_index]-start_year)*12+month-1))
                lon_data_coords = data.coords['lon'].values
                lat_data_coords = data.coords['lat'].values
                
                #Run for first month of the first year because location does not change
                if month == 1 and year_index == 0:
                    
                    #Checks if the latitude input is in range of the Global Climate Model grid
                    if self.lat_desired <= lat_data_coords[0] or self.lat_desired >= lat_data_coords[-1]:
                        raise ValueError("Latitude input is out of range of GCM")
                    
                    #Setting Arbitrary grid points if lon<358.5
                    lon_grid_coords_1 = lon_data_coords[-1]
                    lon_grid_coords_2 = lon_data_coords[0]
                    
                    #Finding the coords of the closest grid points
                    lon_init = lon_data_coords[0]
                    delta_lon = lon_data_coords[1]-lon_init
                    x = math.floor((self.lon_desired - lon_init)/delta_lon)
                
                    lat_init = lat_data_coords[0]
                    delta_lat = (lat_data_coords[-1]-lat_init)/(len(lat_data_coords)-1)
                    y = math.floor((self.lat_desired - lat_init)/delta_lat)
                    
                #Finding distances from the grid points
                if x >= len(lon_data_coords)-1: #Accounting for Longitude Wrapping
                    x = -1
                 
                lon_grid_coords_1 = lon_data_coords[x]
                lon_grid_coords_2 = lon_data_coords[x+1]
                
                if x == -1 and self.lon_desired >= 300: #For data points that are in between [-1] and [0] on the left of lon = 0
                    diff_lon = (lon_grid_coords_2 - (lon_grid_coords_1 - 360))
                    grid_lon_mult = abs(self.lon_desired-lon_grid_coords_1)/diff_lon
                elif x == -1: #For data points that are in between [-1] and [0] on the right of lon = 0
                    diff_lon = (lon_grid_coords_2 - (lon_grid_coords_1 - 360))
                    grid_lon_mult = (1-abs(lon_grid_coords_2-self.lon_desired)/diff_lon)
                else:
                    grid_lon_mult = abs(self.lon_desired-lon_grid_coords_1)/delta_lon
                
                lat_grid_coords_1 = lat_data_coords[y]
                grid_lat_mult = abs(self.lat_desired-lat_grid_coords_1)/delta_lat
                
                if grid_lat_mult > 1:
                    delta_lat = (lat_data_coords[-1]-lat_init)/(len(lat_data_coords)+1)
                    y = math.floor((self.lat_desired - lat_init)/delta_lat)
                    lat_grid_coords_1 = lat_data_coords[y]
                    grid_lat_mult = abs(self.lat_desired-lat_grid_coords_1)/delta_lat
                    
                #Location Error
                if month == 1 and year_index == 0 and self.calculate_error:

                    dist = np.sqrt(0.5) - np.sqrt((0.5-grid_lon_mult)**2 + (0.5-grid_lat_mult)**2)
                    dist_error = (1/np.sqrt(0.5))*dist
                    
                    #Normalization Error Boundaries.Set for the US
                    US_1_lon, US_1_lat = 235, 25
                    US_1_x,US_1_y = math.floor((US_1_lon - lon_init)/delta_lon), math.floor((US_1_lat - lat_init)/delta_lat)
    
                    US_4_lon, US_4_lat = 295, 50
                    US_4_x,US_4_y = math.ceil((US_4_lon - lon_init)/delta_lon), math.ceil((US_4_lat - lat_init)/delta_lat)
                    
                #Establishing temperature values for the 4 surrounding grid points
                temp_1 = float(data.isel(lon=x).isel(lat=y).values)
                temp_2 = float(data.isel(lon=x+1).isel(lat=y).values)
                temp_3 = float(data.isel(lon=x).isel(lat=y+1).values)
                temp_4 = float(data.isel(lon=x+1).isel(lat=y+1).values)     
                
                #Normalization Error
                if self.calculate_error:
                    US_temp_array = data.values[US_1_y:US_4_y,US_1_x:US_4_x] #[rows,cols] -> [y,x]   
                    if np.max(abs(np.diff(US_temp_array))) > total_max_diff:
                        total_max_diff = np.max(abs(np.diff(US_temp_array))) 
                    US_temp_array = np.rot90(US_temp_array)
                    if np.max(abs(np.diff(US_temp_array))) > total_max_diff:
                        total_max_diff = np.max(abs(np.diff(US_temp_array)))
                        
                    #local Normalization Error
                    local_temp_diff = np.max(abs(np.array([temp_1-temp_2,temp_1-temp_3,temp_3-temp_4,temp_2-temp_4])))
                    if local_temp_diff > local_max_diff:
                        local_max_diff = local_temp_diff
                
                temp_inter_1 = temp_1*(1-grid_lon_mult) + temp_2*grid_lon_mult
                temp_inter_2 = temp_3*(1-grid_lon_mult) + temp_4*grid_lon_mult
                temp_des = temp_inter_1*(1-grid_lat_mult) + temp_inter_2*grid_lat_mult
                
                mon_temp[month-1] = temp_des
            
            yearly_avg_temp = sum(mon_temp)/12
            if years[0] == 1850 and years[year_index] > 2014: continue #Skipping years past 2014 for some historical data sets for consistency
            total_object.year_temp_dict[years[year_index]].append(yearly_avg_temp)
        
        if self.calculate_error:
            norm_factor =local_max_diff/total_max_diff
            error = dist_error*norm_factor
        else:
            error = None
        model_object.error = error
        total_object.avg_error_list.append(error)
    
    
    def _calculate_stats(self):
        """
        obj._calculate_stats()
        
        Calculates ensemble means and 95% confidence interval bounds for the 
        total temperature data of each scenario, then performs a fixed point
        regression on all data to produce useable results. The regressed data
        is later outputted as results.
        """
        historical_years = [year for year in range(1850,2015,1)]
        future_years = [year for year in range(2015,2101,1)]
        historical_averaged_temps = [np.mean(self.total_model_data["historical"].year_temp_dict[year]) for year in historical_years]
        
        #Sets fixed point at 2015
        N = len(future_years)
        sigma =np.ones(N)
        sigma[0] = 0.001
        
        def poly_function(x,*p):
            return np.poly1d(p)(x)
        
        #Calculates regressed historical data for baseline calculation
        baseline_polynomial, _ = curve_fit(poly_function, historical_years, historical_averaged_temps, (0, 0, 0))
        baseline = np.poly1d(baseline_polynomial)(self.year_baseline)
        scenario_transition = np.poly1d(baseline_polynomial)(2015)
        
        #Calculates regressed histroical data again, but with the baseline offset
        historical_averaged_temps_baseline = np.array(historical_averaged_temps) - baseline
        historical_polynomial, _ = curve_fit(poly_function, historical_years, historical_averaged_temps_baseline, (0, 0, 0))
        historical_regression = np.poly1d(historical_polynomial)(historical_years)
        self.total_model_data["historical"].delT_list_reg = historical_regression
        
        for year_index in range(len(historical_years)):
            dataset = self.total_model_data["historical"].year_temp_dict[historical_years[year_index]] - baseline
            delT = np.mean(dataset) 
            lower_proj, upper_proj = st.t.interval(alpha=0.95, df=len(dataset)-1, loc=np.mean(dataset), scale=st.sem(dataset))
            CI = upper_proj - delT
            self.total_model_data["historical"].CI_list[year_index] = CI
            self.total_model_data["historical"].delT_list[year_index] = delT
        
        for scenario in self.scenario_list:
            if scenario == "historical": continue
            averaged_temps = [np.mean(self.total_model_data[scenario].year_temp_dict[year]) for year in future_years]
            averaged_temps[0] = scenario_transition
            averaged_temps_delta = np.array(averaged_temps) - baseline
            polynomial, _ = curve_fit(poly_function, future_years, averaged_temps_delta, (0, 0, 0), sigma=sigma)
            regression = np.poly1d(polynomial)(future_years)
            self.total_model_data[scenario].delT_list_reg = regression
            
            for year_index in range(len(future_years)):
                dataset = self.total_model_data[scenario].year_temp_dict[future_years[year_index]] - baseline
                delT = np.mean(dataset)
                lower_proj, upper_proj = st.t.interval(alpha=0.95, df=len(dataset)-1, loc=np.mean(dataset), scale=st.sem(dataset))
                CI = upper_proj - delT
                self.total_model_data[scenario].CI_list[year_index] = CI
                self.total_model_data[scenario].delT_list[year_index] = delT
                

    def results1(self,scatter_display=[False,False,False,False,False,False],regression_display=[True,True,True,True,True,True],CI_display=[True,False,True,False,True,False]):
        """
        obj.results1(scatter_display=[False,False,False,False,False,False],
                    regression_display=[True,True,True,True,True,True],
                    CI_display=[True,False,True,False,True,False])
        
        Parameters
        ----------
        
        scatter_display : bool : list : optional : Default = [False,False,False,False,False,False]
            A list of 6 boolean values used for selecting which scatter data to plot.
            The list items correspond to: historical, SSP119, SSP126, SSP245, SSP370, SSP585.
            
        regression_display : bool : list : optional : Default = [True,True,True,True,True,True]
            A list of 6 boolean values used for selecting which regressions to plot.
            The list items correspond to: historical, SSP119, SSP126, SSP245, SSP370, SSP585.
            
        CI_display : bool : list : optional : Default = [True,False,True,False,True,False]
            A list of 6 boolean values used for selecting which Confidence intervals to plot.
            The list items correspond to: historical, SSP119, SSP126, SSP245, SSP370, SSP585.            
        
        Returns
        -------
        
        None - plots the temperature change values for the scenarios for the years
        1950-2100.
        """
        
        historical_years = [year for year in range(1850,2015,1)]
        future_years = [year for year in range(2015,2101,1)]
        
        plt.rcParams.update({'font.size': 22})
        fig,ax = plt.subplots()
        
        #years, colors, and labels organized into dictionaries for succinctness
        year_dict = {"historical":historical_years,"SSP119":future_years,"SSP126":future_years,"SSP245":future_years,"SSP370":future_years,"SSP585":future_years}
        color_dict = {"historical":"k","SSP119":"dodgerblue","SSP126":"navy","SSP245":"gold","SSP370":"red","SSP585":"maroon"}
        scatter_label_dict = {"historical":"Averaged Historical Data","SSP119":"Averaged SSP1-1.9 Data","SSP126":"Averaged SSP1-2.6 Data",
                              "SSP245":"Averaged SSP2-4.5 Data","SSP370":"Averaged SSP3.70 Data","SSP585":"Averaged SSP5-8.5 Data"}
        regression_label_dict = {"historical":"Historical Regression","SSP119":"SSP1-1.9 Regression","SSP126":"SSP1-2.6 Regression",
                                 "SSP245":"SSP2-4.5 Regression","SSP370":"SSP3.70 Regression","SSP585":"SSP5-8.5 Regression"}
        CI_label_dict = {"historical":"Historical Data 95% CI","SSP119":"SSP1-1.9 Data 95% CI","SSP126":"SSP1-2.6 Data 95% CI",
                         "SSP245":"SSP2-4.5 Data 95% CI","SSP370":"SSP3.70 Data 95% CI","SSP585":"SSP5-8.5 Data 95% CI"}

        for scenario in self.scenario_list:
            index = self.scenario_list.index(scenario)
            if scatter_display[index]:
                ax.scatter(year_dict[scenario],self.total_model_data[scenario].delT_list,c=color_dict[scenario],label=scatter_label_dict[scenario])
            if regression_display[index]:
                ax.plot(year_dict[scenario],self.total_model_data[scenario].delT_list_reg,color_dict[scenario],linewidth=3,label=regression_label_dict[scenario])
            if CI_display[index]:
                ax.fill_between(year_dict[scenario],(self.total_model_data[scenario].delT_list-self.total_model_data[scenario].CI_list),
                                (self.total_model_data[scenario].delT_list+self.total_model_data[scenario].CI_list),color=color_dict[scenario],alpha=.2,label=CI_label_dict[scenario])
            
        plot_box = ax.get_position()
        ax.set_position([plot_box.x0, plot_box.y0,plot_box.width*0.8,plot_box.height])
        ax.legend(loc='center left', bbox_to_anchor=(1, 0.5))
        plt.xlabel("Year",fontdict={'fontsize': 26})
        plt.ylabel("°C Change",fontdict={'fontsize': 26})
        plt.title("Surface Temperature Change at ({0:.5f}, {1:.5f}) Relative to {2:.0f}".format(self.lat_desired,self.lon_desired,self.year_baseline),y=1.02,x=0.7,fontdict={'fontsize': 30})
        plt.xlim(1950,2100)
        plt.grid()
        fig.set_size_inches(18.5, 10.5)
        plt.show()
        
    
    def results2(self,desired_scenario,resolution='low'):
        """
        obj.results2(desired_scenario,
                     resolution='low')
        
        Parameters
        ----------
        
        desired_scenario : str 
            A str that selectes the future scenario heat map that will be calculated and
            displayed. The possible strings are: "SSP119","SSP126","SSP245","SSP370","SSP585".
            Any other strings will return an error.
            
        resolution : str : optional : Default = 'low'
            A string that selectes the resolution of the heat map plotted. The default 
            setting is low so the simulation can be run quicker. Higher qualities can
            take much longer to calculate and display. All possible resolutions are:
            "low","medium","high","extreme". "test" is very inaccurate, only used for
            unittesting.
            
        Returns
        -------
        
        None - plots the heat map of the change of temperature from a baseline year
        to a future year for the United States for a selected scenario.
        """
        resolution_dict = {"test":[2,2],"low":[60,27],"medium":[120,54],"high":[240,108],"extreme":[480,216]}
 
        lon_array = np.linspace(360-125,360-66,num=resolution_dict[resolution][0])
        lat_array = np.linspace(24,50,num=resolution_dict[resolution][1])
        folder_path = "CMIP6_Data_Files" 
        
        result_arrays=[]
        for scenario in ['historical',desired_scenario]:
            year = self.year_baseline if scenario == 'historical' else self.year_desired
    
            
            model_guide_wb = openpyxl.load_workbook(os.path.join(self.dirname,self.model_guide))
            model_guide_ws = model_guide_wb[scenario]
        
            model_list = []
            for row in range(2,model_guide_ws.max_row+1):
                name = model_guide_ws.cell(row,1).value
                if name not in model_list:
                    model_list.append(name)
                    
            pool = mp.Pool(mp.cpu_count()-1)
            temp_array_list = []
            for model in model_list:
                temp_array_list.append(pool.apply_async(self._results2_calc,(model_guide_ws,folder_path,scenario,model,lat_array,lon_array,year)))
            pool.close()
            pool.join()
    
            for index in range(len(temp_array_list)):
                temp_array_list[index] = temp_array_list[index].get()
        
            array = temp_array_list[0]
            if len(model_list) > 1:
                for index in range(1,len(model_list)):
                    array  += temp_array_list[index]
                array = array/len(model_list)
            result_arrays.append(array)
        
        baseline = xr.DataArray(data=result_arrays[0],dims=["lat", "lon"],coords=[lat_array,lon_array])
        projection = xr.DataArray(data=result_arrays[1],dims=["lat", "lon"],coords=[lat_array,lon_array])
        delta_temp = projection - baseline

        value_high = np.max(delta_temp)
        value_low = abs(np.min(delta_temp))
        limit = np.max([math.ceil(value_high),math.ceil(value_low)])

        extent = [-125, -66, 24, 46]
        plt.rcParams.update({'font.size': 12})
        
        if self.use_cartopy:
            plt.figure(figsize=(16, 6))
            ax = plt.subplot(1, 1, 1, projection=ccrs.PlateCarree())
            delta_temp.plot.pcolormesh(ax=ax,cmap='coolwarm',vmin=-limit,vmax=limit)
            ax.gridlines(draw_labels=True)
            ax.set_extent(extent)
            ax.coastlines()
            ax.add_feature(cf.BORDERS)
            ax.add_feature(cf.STATES)
            plt.title(f"US Temperature Change from {self.year_baseline} to {self.year_desired} According to {desired_scenario}",x=0.55,fontdict={'fontsize': 16})
        else:
            logger.warning("Cannot display US borders because cartopy is not properly installed")
            fig = plt.figure()
            ax = plt.axes([0,0,2,1])
            c = ax.pcolormesh(lon_array, lat_array, delta_temp, cmap='coolwarm', vmin=-limit, vmax=limit)
            ax.set_title('pcolormesh')
            fig.colorbar(c, ax=ax)
            plt.ylabel("Latitude [Degrees North]")
            plt.xlabel("Longitude [Degrees East]")
            plt.xlim(235,294)
            plt.ylim(24,50)
            plt.grid()
            plt.title(f"US Temperature Change from {self.year_baseline} to {self.year_desired} According to {desired_scenario}")
            
        plt.show()

            
    def _results2_calc(self,model_guide_ws,folder_path,scenario,model,lat_array,lon_array,year):
        """
        obj._results2_calc(model_guide_ws,
                           folder_path,
                           scenario,
                           model,
                           lat_array,
                           lon_array,
                           year)
        
        Performs all of the calculations associated with plotting the US heat maps
        in the function call obj.results2(desired_scenario, resolution). First, it
        compiles the data used, then computes the temperature change grid array, 
        and finally returns that data to results2.
        """
        data_list = []
        for row in range(2,model_guide_ws.max_row+1):
            model_name = model_guide_ws.cell(row,1).value
            link = model_guide_ws.cell(row,2).value
            file_path_list = link.split("/")
            file_name = file_path_list[-1] 
            path = os.path.join(self.dirname,folder_path,scenario,file_name)
            if model_name == model:
                Model = self.Model_object(model_name,path)
                file_count = model_guide_ws.cell(row,3).value
                file_num = model_guide_ws.cell(row,4).value
                if file_count == file_num:
                    if data_list != []:
                        Complete_model = Model
                        data_list.append(Model.data)
                        setattr(Complete_model,'data', xr.concat(data_list,dim='time'))
                    else:
                        Complete_model = Model
                else:
                    data_list.append(Model.data)
        
        Complete_model.set_data()
        temperature_baseline = np.zeros((len(lat_array),len(lon_array)))
        for i in range(len(lat_array)):
            for j in range(len(lon_array)):
                lat = lat_array[i]
                lon = lon_array[j]
            
                mon_temp = [0]*12
                for month in range(1,13):
                    data = Complete_model.data.isel(time=int((year-Complete_model.start_year)*12+month-1))
                    lon_data_coords = data.coords['lon'].values
                    lat_data_coords = data.coords['lat'].values

                    if month == 1:
                        lon_init = lon_data_coords[0]
                        delta_lon = lon_data_coords[1]-lon_init
                        x = math.floor((lon - lon_init)/delta_lon)
                    
                        lat_init = lat_data_coords[0]
                        delta_lat = (lat_data_coords[-1]-lat_init)/(len(lat_data_coords)-1)
                        y = math.floor((lat - lat_init)/delta_lat)
   
                    lon_grid_coords_1 = lon_data_coords[x]
                    grid_lon_mult = abs(lon-lon_grid_coords_1)/delta_lon
                    lat_grid_coords_1 = lat_data_coords[y]
                    grid_lat_mult = abs(lat-lat_grid_coords_1)/delta_lat
                    
                    if grid_lat_mult > 1:
                        delta_lat = (lat_data_coords[-1]-lat_init)/(len(lat_data_coords)+1)
                        y = math.floor((lat - lat_init)/delta_lat)
                        lat_grid_coords_1 = lat_data_coords[y]
                        grid_lat_mult = abs(lat-lat_grid_coords_1)/delta_lat
                        
                    #Establishing temperature values for the 4 surrounding grid points
                    temp_1 = float(data.isel(lon=x).isel(lat=y).values)
                    temp_2 = float(data.isel(lon=x+1).isel(lat=y).values)
                    temp_3 = float(data.isel(lon=x).isel(lat=y+1).values)
                    temp_4 = float(data.isel(lon=x+1).isel(lat=y+1).values)     
                
                    temp_inter_1 = temp_1*(1-grid_lon_mult) + temp_2*grid_lon_mult
                    temp_inter_2 = temp_3*(1-grid_lon_mult) + temp_4*grid_lon_mult
                    temp_des = temp_inter_1*(1-grid_lat_mult) + temp_inter_2*grid_lat_mult
                    
                    mon_temp[month-1] = temp_des
                yearly_avg_temp = sum(mon_temp)/12
                temperature_baseline[i][j] = yearly_avg_temp
                
        return temperature_baseline

        

if __name__ == '__main__':
    #Test Code: example ran in examples/CMIP_data
    start_time = time.time()
    obj = CMIP_Data(lat_desired = 35.0433,
                    lon_desired = -106.6129,
                    year_baseline = 2014,
                    year_desired = 2050,
                    file_path = os.path.abspath(os.getcwd()),
                    model_guide = "Models_Used_Simplified.xlsx",
                    calculate_error=True,
                    world_map=True,
                    display_logging=False)
    obj.results1(scatter_display=[True,True,True,True,True,True])
    #obj.results2(desired_scenario = "SSP119",resolution = "high")

    #print("Program run time: {0:.3f} seconds".format(time.time() - start_time))


"""
Cities chosen based on climate zones
Miami lat: 25.7881 Lon: -80.31694
Houston lat: 29.9844 lon: -95.36072
Phooenix lat: 33.4278 lon: -112.0037
Atlanta lat: 33.6297 lon: -84.44224
Los Angeles lat: 33.9382 lon: -118.3866
Las Vegas lat: 36.2121 lon: -115.1939
San Francisco lat: 37.6196 lon: -122.3656
Baltimore lat: 39.1733 lon: -76.68408
Albuquerque lat: 35.0433 lon: -106.6129
Seattle lat: 47.4447 lon: -122.3144
Chicago lat: 41.7841 lon: -87.75514
Denver lat: 39.8466 lon: -104.6562
Minneapolis lat: 44.8852 lon: -93.23133
Helena lat: 46.6044 lon: -111.9892
Duluth lat: 46.8436 lon: -92.18658
Fairbanks lat: 64.8031 lon: -147.8761
"""